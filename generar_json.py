"""
=============================================================
 generar_json.py  –  RDC I+D PMO
=============================================================
 QUÉ HACE ESTE SCRIPT:
   Lee el archivo Excel de seguimiento de proyectos I+D
   y genera automáticamente el archivo proyectos.json
   que usa el dashboard HTML.

 CÓMO USARLO (cada mes):
   1. Abrir la terminal (cmd o PowerShell en Windows)
   2. Ir a la carpeta donde está este archivo:
        cd "C:\\Users\\user\\OneDrive - REFINERIA DEL CENTRO S.A\\CLAUDE\\PROYECTOS I+D\\GESTION DE PROYECTOS I+D\\Instrucciones"
   3. Ejecutar:
        python generar_json.py
   4. El archivo proyectos.json se actualiza solo.
   5. Abrir el dashboard en el navegador. Listo.

 REQUISITOS:
   pip install openpyxl
=============================================================
"""

import json
import os
from datetime import datetime, date
import warnings
warnings.filterwarnings("ignore")

# ── CONFIGURACIÓN ──────────────────────────────────────────
# Ajustá esta ruta al nombre exacto de tu archivo Excel
EXCEL_FILE = "../0326 - RDC+D - REP PROY I+D (v3)-Note-IDiaz.xlsm"

# Nombre del archivo JSON que se va a generar
JSON_OUTPUT = "proyectos.json"

# Hoja principal del reporte
SHEET_NAME = "REPORTE"

# Fila donde empiezan los datos (fila 6 en Excel = índice 5 en Python)
FILA_INICIO = 6

# Fila máxima a leer (ajustar si se agregan proyectos)
FILA_FIN = 47

# Fecha de corte (hoy)
FECHA_CORTE = date.today().isoformat()
# ──────────────────────────────────────────────────────────


def fecha_a_str(valor):
    """Convierte una fecha de Excel a string 'YYYY-MM-DD', o devuelve None."""
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, str) and len(valor) >= 10:
        return valor[:10]
    return None


def calcular_desvio_meses(fecha_plan, fecha_corte_str):
    """
    Calcula cuántos meses de desvío tiene un hito:
      - Negativo: el hito aún NO venció (está en el futuro)
      - Positivo: el hito YA venció (está atrasado)
      - Cero:     el hito fue exactamente hoy
    """
    if not fecha_plan:
        return None
    try:
        plan = datetime.strptime(fecha_plan, "%Y-%m-%d").date()
        corte = datetime.strptime(fecha_corte_str, "%Y-%m-%d").date()
        delta_dias = (corte - plan).days
        return round(delta_dias / 30.44, 2)  # meses = días / 30.44
    except Exception:
        return None


def calcular_regla(desvio):
    """
    Convierte el desvío en meses a una regla de semáforo:
      1.00 = verde (en tiempo o futuro)
      0.75 = amarillo (hasta 1 mes de atraso)
      0.50 = naranja (1-3 meses de atraso)
      0.25 = rojo (más de 3 meses de atraso)
    """
    if desvio is None:
        return 1.0
    if desvio <= 0:
        return 1.0
    elif desvio <= 1:
        return 0.75
    elif desvio <= 3:
        return 0.50
    else:
        return 0.25


def valor_o_none(v):
    """Limpia valores vacíos o de error de Excel."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "N/A", "–", "-"):
            return None
    return v


def leer_excel_y_generar_json():
    """Función principal: lee el Excel y genera el JSON."""

    # ── 1. Verificar que existe el Excel ──────────────────
    ruta_script = os.path.dirname(os.path.abspath(__file__))
    ruta_excel = os.path.join(ruta_script, EXCEL_FILE)

    if not os.path.exists(ruta_excel):
        print(f"❌ ERROR: No se encontró el archivo Excel:")
        print(f"   {ruta_excel}")
        print(f"\n   Verificá que el nombre del archivo sea exactamente:")
        print(f"   {EXCEL_FILE}")
        return

    print(f"📂 Leyendo Excel: {os.path.basename(ruta_excel)}")

    # ── 2. Abrir el Excel ─────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta_excel, data_only=True, keep_vba=False)
    except ImportError:
        print("❌ ERROR: Falta instalar openpyxl.")
        print("   Ejecutá: pip install openpyxl")
        return
    except Exception as e:
        print(f"❌ ERROR al abrir el Excel: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ ERROR: No se encontró la hoja '{SHEET_NAME}'")
        print(f"   Hojas disponibles: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    print(f"✅ Hoja '{SHEET_NAME}' encontrada.")

    # ── 3. Leer filas de proyectos ────────────────────────
    #
    # MAPA DE COLUMNAS (basado en el archivo real):
    #   Col A (0)  = Tipo de fila ('KPI' o vacío)
    #   Col B (1)  = ID del proyecto
    #   Col C (2)  = Nombre del proyecto
    #   Col D (3)  = UN
    #   Col E (4)  = Planta
    #   Col F (5)  = Sector
    #   Col G (6)  = PM (Responsable)
    #   Col H (7)  = Prioridad
    #   Col I (8)  = Descripción
    #   Col K (10) = Fecha inicio
    #   Col M (12) = GTP1 Plan
    #   Col N (13) = GTP1 Real
    #   Col O (14) = Hito actual (estado)
    #   Col P (15) = Desvío GTP1 (meses)
    #   Col Q (16) = Avance entregables
    #   Col T (19) = GTP2 Plan
    #   Col U (20) = GTP2 Real
    #   Col V (21) = Desvío GTP2 (meses)
    #   Col AX(49) = Score año en curso

    proyectos = []
    kpi_actual = "SIN KPI"
    proyectos_leidos = 0
    filas_saltadas = 0

    for fila_num in range(FILA_INICIO, FILA_FIN + 1):
        fila = ws[fila_num]

        # Obtener valores de la fila (índice base 0)
        def cel(col_idx):
            """Devuelve el valor limpio de una celda por índice (0=A, 1=B...)"""
            v = fila[col_idx].value if col_idx < len(fila) else None
            return valor_o_none(v)

        tipo_fila = cel(0)   # Columna A
        id_proy   = cel(1)   # Columna B
        nombre    = cel(2)   # Columna C

        # ── Detectar filas KPI (encabezados de categoría) ──
        if str(tipo_fila).strip().upper() == "KPI":
            kpi_actual = str(nombre).strip() if nombre else kpi_actual
            print(f"   📌 KPI encontrado: {kpi_actual}")
            continue

        # ── Saltar filas sin nombre de proyecto ──
        if not nombre:
            filas_saltadas += 1
            continue

        # ── Saltar filas de totales/fórmulas ──
        if id_proy and str(id_proy).strip().upper() == "KPI":
            continue

        # ── Leer campos del proyecto ──
        sector   = cel(5)   # Columna F
        pm       = cel(6)   # Columna G
        prioridad = cel(7)  # Columna H
        estado   = str(cel(14) or "").strip().upper()  # Columna O

        # Normalizar estado
        if "EJECUC" in estado:
            estado = "EJECUCIÓN"
        elif "EXPLOR" in estado:
            estado = "EXPLORACIÓN"
        elif "FINALIZ" in estado:
            estado = "FINALIZADO"
        elif "DETENI" in estado:
            estado = "DETENIDO"
        else:
            estado = estado or "DESCONOCIDO"

        # Fechas
        g1p = fecha_a_str(cel(12))   # GTP1 Plan (col M)
        g1r = fecha_a_str(cel(13))   # GTP1 Real (col N)
        g2p = fecha_a_str(cel(19))   # GTP2 Plan (col T)
        g2r = fecha_a_str(cel(20))   # GTP2 Real (col U)

        # Desvíos (tomamos del Excel si están, sino calculamos)
        d1_excel = cel(15)  # Col P = Desvío GTP1
        d2_excel = cel(21)  # Col V = Desvío GTP2

        if d1_excel is not None and isinstance(d1_excel, (int, float)):
            d1 = round(float(d1_excel), 2)
        else:
            # Calcular desvío GTP1: si cerrado usar g1r, sino usar g1p vs hoy
            fecha_ref = g1r if g1r else g1p
            d1 = calcular_desvio_meses(fecha_ref, FECHA_CORTE)

        if d2_excel is not None and isinstance(d2_excel, (int, float)):
            d2 = round(float(d2_excel), 2)
        else:
            fecha_ref = g2r if g2r else g2p
            d2 = calcular_desvio_meses(fecha_ref, FECHA_CORTE)

        # Avance de entregables
        av_raw = cel(16)  # Col Q
        av = round(float(av_raw), 2) if isinstance(av_raw, (int, float)) else 1.0

        # Score anual
        sc_raw = cel(49)  # Col AX
        sc = round(float(sc_raw), 3) if isinstance(sc_raw, (int, float)) else None

        # Regla de semáforo
        rl = calcular_regla(d1)

        # Normalizar ID
        id_str = str(id_proy).strip() if id_proy else "–"

        proyecto = {
            "id":  id_str,
            "n":   str(nombre).strip(),
            "k":   kpi_actual,
            "s":   str(sector).strip() if sector else "–",
            "pm":  str(pm).strip() if pm else "–",
            "pr":  str(prioridad).strip() if prioridad else "–",
            "est": estado,
            "g1p": g1p,
            "g1r": g1r,
            "g2p": g2p,
            "g2r": g2r,
            "d1":  d1,
            "d2":  d2,
            "rl":  rl,
            "av":  av,
            "sc":  sc
        }

        proyectos.append(proyecto)
        proyectos_leidos += 1

    print(f"\n📊 Proyectos encontrados: {proyectos_leidos}")
    print(f"   Filas saltadas (vacías/totales): {filas_saltadas}")

    # ── 4. Calcular resumen ───────────────────────────────
    por_estado = {}
    por_pm = {}
    con_desvio = 0

    for p in proyectos:
        # Contar por estado
        e = p["est"]
        por_estado[e] = por_estado.get(e, 0) + 1

        # Contar por PM
        pm_name = p["pm"]
        por_pm[pm_name] = por_pm.get(pm_name, 0) + 1

        # Contar desvíos
        if p["d1"] is not None and p["d1"] > 0:
            con_desvio += 1

    # Calcular score GPO general (promedio de scores no nulos)
    scores = [p["sc"] for p in proyectos if p["sc"] is not None]
    score_gpo = round(sum(scores) / len(scores), 3) if scores else 0

    # ── 5. Armar el JSON final ────────────────────────────
    output = {
        "_metadata": {
            "descripcion": "Datos del portfolio de proyectos I+D – RDC",
            "fuente": os.path.basename(ruta_excel),
            "corte": FECHA_CORTE,
            "version": "1.0",
            "generado_por": "generar_json.py",
            "generado_en": datetime.now().isoformat(timespec="seconds"),
            "total_proyectos": len(proyectos),
            "score_gpo": score_gpo,
            "resumen": {
                "por_estado": por_estado,
                "por_pm": por_pm,
                "con_desvio": con_desvio
            }
        },
        "proyectos": proyectos
    }

    # ── 6. Guardar el JSON ────────────────────────────────
    ruta_json = os.path.join(ruta_script, JSON_OUTPUT)

    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ JSON generado exitosamente:")
    print(f"   📄 {ruta_json}")
    print(f"   📦 Tamaño: {os.path.getsize(ruta_json):,} bytes")
    print(f"\n📈 RESUMEN DEL PORTFOLIO:")
    for estado, cant in sorted(por_estado.items()):
        print(f"   {estado}: {cant} proyectos")
    print(f"   Con desvío: {con_desvio}")
    print(f"   Score GPO aprox: {score_gpo*100:.1f}%")
    print(f"\n🎉 Listo. Ahora abrí el dashboard en tu navegador.")


# ── PUNTO DE ENTRADA ──────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  RDC I+D – Generador de datos para PMO Dashboard")
    print("=" * 55)
    leer_excel_y_generar_json()
    print("=" * 55)
